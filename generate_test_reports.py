"""
Generate four separate Excel test report artifacts:
  1. Selenium Test Report  (400 PASS)
  2. Appium Test Report    (400 PASS)
  3. Load Test Report      (400 PASS)
  4. Vulnerability Test Report (400 PASS)

Each workbook contains:
  • Summary sheet  – high-level metrics
  • Detailed Report sheet – 400 rows of mock test-case data

All data is mock / synthetic.  Every test case has status = PASS.
"""

import os
import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────
NUM_CASES = 350
EXECUTION_DATETIME = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# Styling constants
HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUMMARY_LABEL_FONT = Font(name="Calibri", bold=True, size=11)
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(name="Calibri", bold=True, color="006100")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# ──────────────────────────────────────────────
# Domain-specific mock data pools
# ──────────────────────────────────────────────

SELENIUM_MODULES = [
    "Login & Authentication", "User Registration", "Dashboard",
    "Patient Profile", "Search & Filter", "Reports & Analytics",
    "Settings & Preferences", "Navigation & Routing", "Form Validation",
    "Data Table & Pagination", "File Upload/Download", "Notification System",
    "Session Management", "Role-Based Access", "Audit Trail",
    "ECG Viewer", "Appointment Scheduler", "Prescription Module",
    "Billing & Invoice", "Admin Panel",
]

SELENIUM_ACTIONS = [
    ("Verify {module} page loads correctly", "Page loads within 3 seconds with all elements visible", "Page loaded in {t}s, all elements rendered"),
    ("Validate {module} form submission", "Form submits successfully and confirmation displayed", "Form submitted, confirmation message shown"),
    ("Check {module} input field validation", "Invalid input shows error message", "Error message displayed for invalid input"),
    ("Test {module} button click action", "Button triggers expected action", "Action triggered successfully on click"),
    ("Verify {module} breadcrumb navigation", "Breadcrumb shows correct hierarchy", "Breadcrumb navigation path is correct"),
    ("Validate {module} dropdown selection", "Dropdown populates and selection is retained", "Dropdown selection retained after page action"),
    ("Test {module} responsive layout", "Layout adapts to viewport size", "Layout correctly adapted to {w}px width"),
    ("Verify {module} modal dialog opens and closes", "Modal opens on trigger and closes on dismiss", "Modal opened and closed successfully"),
    ("Check {module} table sort functionality", "Table sorts by selected column", "Table sorted correctly by column"),
    ("Validate {module} pagination controls", "Pagination navigates between pages", "Page {p} loaded with correct records"),
    ("Test {module} tooltip display", "Tooltip appears on hover", "Tooltip displayed with correct text"),
    ("Verify {module} error state rendering", "Error state UI displays properly", "Error state rendered with fallback UI"),
    ("Check {module} loading spinner display", "Spinner shown during data fetch", "Loading spinner displayed during API call"),
    ("Validate {module} date picker interaction", "Date picker opens and date is selectable", "Date selected and populated in field"),
    ("Test {module} tab switching", "Tabs switch and display correct content", "Tab content switched correctly"),
    ("Verify {module} link redirection", "Link redirects to correct URL", "Redirected to expected URL"),
    ("Validate {module} checkbox toggle", "Checkbox toggles state correctly", "Checkbox state toggled as expected"),
    ("Test {module} auto-complete suggestion", "Suggestions appear after typing 3 chars", "Suggestions displayed after typing"),
    ("Verify {module} scroll behavior", "Page scrolls smoothly to section", "Smooth scroll to target section confirmed"),
    ("Check {module} logout flow", "User is logged out and redirected to login", "Logout successful, redirected to login page"),
]

APPIUM_MODULES = [
    "Mobile Login", "Biometric Auth", "Push Notifications",
    "Mobile Dashboard", "Patient Vitals", "Heart Rate Monitor",
    "ECG Display", "Medication Tracker", "Appointment Booking",
    "Profile Management", "Offline Mode", "Sync & Data Refresh",
    "Camera Integration", "QR Code Scanner", "Gesture Controls",
    "Deep Linking", "App Settings", "Onboarding Flow",
    "Mobile Search", "Alert & Reminders",
]

APPIUM_ACTIONS = [
    ("Verify {module} screen renders on launch", "Screen renders within 2 seconds", "Screen rendered in {t}s"),
    ("Validate {module} tap interaction", "Tap triggers expected navigation", "Navigation triggered on tap"),
    ("Test {module} swipe gesture", "Swipe navigates to next card/page", "Swipe gesture handled correctly"),
    ("Check {module} text input on mobile keyboard", "Keyboard opens and text is entered", "Text entered via mobile keyboard successfully"),
    ("Verify {module} back button behavior", "Back button returns to previous screen", "Returned to previous screen"),
    ("Validate {module} pull-to-refresh", "Pull-to-refresh updates content", "Content refreshed on pull gesture"),
    ("Test {module} landscape/portrait rotation", "UI adjusts on orientation change", "UI adjusted to {orient} mode"),
    ("Check {module} scroll performance", "Scroll is smooth with no jank", "Scroll FPS: {fps}, no jank detected"),
    ("Verify {module} toast message display", "Toast message appears on action", "Toast message displayed correctly"),
    ("Validate {module} permission dialog", "Permission dialog shows and handles response", "Permission dialog handled successfully"),
    ("Test {module} app background/foreground", "App resumes state after backgrounding", "App state restored from background"),
    ("Check {module} local storage persistence", "Data persists after app restart", "Data found in local storage after restart"),
    ("Verify {module} network error handling", "Graceful error shown on no network", "Network error message displayed"),
    ("Validate {module} loading skeleton", "Skeleton UI shown during data load", "Skeleton placeholder rendered"),
    ("Test {module} button disabled state", "Button is disabled when condition met", "Button disabled state verified"),
    ("Check {module} animation smoothness", "Animations run at 60fps", "Animation frame rate: {fps}fps"),
    ("Verify {module} accessibility labels", "Accessibility labels present on elements", "All accessibility labels verified"),
    ("Validate {module} font scaling", "Text scales with system font size", "Font scaled correctly to {scale}x"),
    ("Test {module} dark mode rendering", "UI renders correctly in dark mode", "Dark mode theme applied successfully"),
    ("Check {module} deep link resolution", "Deep link opens correct screen", "Deep link resolved to target screen"),
]

LOAD_MODULES = [
    "API Gateway", "Auth Service", "Patient Records API",
    "ECG Data Stream", "Vitals Ingestion", "Report Generation",
    "Search Service", "Notification Dispatch", "Billing API",
    "Appointment Service", "Prescription API", "Image Upload Service",
    "WebSocket Server", "Cache Layer", "Database Read Pool",
    "Database Write Pool", "Message Queue", "Audit Logging Service",
    "Analytics Pipeline", "CDN & Static Assets",
]

LOAD_ACTIONS = [
    ("Stress test {module} with {n} concurrent users", "Response time < {rts}ms at {n} users", "Avg response: {rt}ms, P99: {p99}ms"),
    ("Ramp-up load test on {module} from 10 to {n} users", "No errors during ramp-up", "0 errors across ramp-up; peak {rt}ms"),
    ("Spike test {module} with sudden {n} user burst", "System recovers within 5 seconds", "Recovery in {rec}s, no dropped requests"),
    ("Soak test {module} over 30-minute window", "Memory stable, no leaks", "Memory delta: {mem}MB over 30 min"),
    ("Throughput test {module} at {tps} req/s", "Sustained throughput without errors", "Sustained {tps} req/s, error rate 0%"),
    ("Latency percentile test {module}", "P50 < {p50}ms, P95 < {p95}ms, P99 < {p99}ms", "P50={p50r}ms, P95={p95r}ms, P99={p99r}ms"),
    ("Connection pool saturation test {module}", "Pool handles {n} connections", "Pool managed {n} connections, 0 timeouts"),
    ("Payload size test {module} with {sz}KB body", "Handles large payloads without timeout", "Processed {sz}KB payload in {rt}ms"),
    ("Rate limiter validation on {module}", "429 returned after {rl} req/min threshold", "Rate limit enforced at {rl} req/min"),
    ("Concurrent write test on {module}", "All writes committed without conflict", "{n} concurrent writes, 0 conflicts"),
    ("Cache hit ratio test {module}", "Cache hit ratio > 90%", "Cache hit ratio: {chr}%"),
    ("Failover test {module} under load", "Failover completes within 3s", "Failover completed in {fo}s"),
    ("Compression efficiency test {module}", "gzip reduces payload by > 60%", "Compression ratio: {comp}%"),
    ("Keep-alive connection reuse test {module}", "Connections reused > 80%", "Connection reuse: {reuse}%"),
    ("Database query performance under load {module}", "Query time < {qt}ms under {n} users", "Avg query: {qtr}ms under {n} users"),
    ("Auto-scaling trigger test {module}", "New instance spins up at {th}% CPU", "Auto-scale triggered at {thr}% CPU"),
    ("Graceful degradation test {module}", "Non-critical features disabled under load", "Degradation policy activated, core stable"),
    ("Batch processing throughput {module}", "Processes {batch} records/min", "Processed {batch} records/min, 0 failures"),
    ("Concurrent read test {module}", "{n} reads/s with consistent data", "{n} reads/s, data consistency 100%"),
    ("Timeout behavior test {module}", "Timeout at {to}s returns proper error", "Timeout at {to}s, error code 504 returned"),
]

VULN_MODULES = [
    "Authentication Endpoint", "Session Management", "Password Policy",
    "Input Validation Layer", "API Authorization", "File Upload Handler",
    "CORS Configuration", "CSRF Protection", "XSS Prevention",
    "SQL Injection Defense", "Rate Limiting", "Encryption at Rest",
    "TLS/SSL Configuration", "Token Management", "OAuth2 Flow",
    "Cookie Security", "HTTP Security Headers", "Error Handling",
    "Dependency Scanner", "Secrets Management",
]

VULN_ACTIONS = [
    ("SQL injection test on {module}", "No SQL injection vulnerability", "Input sanitized; no injection vector found"),
    ("XSS reflected payload test on {module}", "No reflected XSS", "Output encoded; reflected XSS not possible"),
    ("XSS stored payload test on {module}", "No stored XSS vulnerability", "Stored content sanitized, XSS mitigated"),
    ("CSRF token validation on {module}", "CSRF token required and validated", "CSRF token validated on all state-changing requests"),
    ("Directory traversal test on {module}", "Path traversal blocked", "Path normalization prevents traversal"),
    ("Broken authentication check on {module}", "Auth enforced on all endpoints", "Unauthenticated requests returned 401"),
    ("Sensitive data exposure test on {module}", "No sensitive data in response", "PII fields masked in API response"),
    ("Security misconfiguration scan on {module}", "No misconfiguration found", "Configuration hardened, no issues detected"),
    ("Insecure deserialization test on {module}", "Deserialization is safe", "Deserialization uses allowlist, no RCE vector"),
    ("Broken access control test on {module}", "Horizontal privilege escalation blocked", "Access control enforced per user role"),
    ("XML external entity (XXE) test on {module}", "XXE processing disabled", "XML parser configured to disallow DTDs"),
    ("Server-side request forgery (SSRF) test on {module}", "SSRF vectors blocked", "URL allowlist enforced, SSRF mitigated"),
    ("Open redirect test on {module}", "No open redirect vulnerability", "Redirect URLs validated against allowlist"),
    ("Clickjacking protection test on {module}", "X-Frame-Options header present", "X-Frame-Options: DENY header set"),
    ("Content-Type header validation on {module}", "Strict Content-Type enforcement", "Content-Type validated on all endpoints"),
    ("HTTP Strict Transport Security test on {module}", "HSTS header present", "HSTS header: max-age=31536000; includeSubDomains"),
    ("Cookie security flags test on {module}", "Secure, HttpOnly, SameSite flags set", "All cookies have Secure, HttpOnly, SameSite=Strict"),
    ("Password brute-force protection on {module}", "Account lockout after 5 attempts", "Account locked after 5 failed attempts"),
    ("JWT token validation test on {module}", "JWT signature verified, expiry enforced", "JWT signature valid, expired tokens rejected"),
    ("Dependency vulnerability scan on {module}", "No known CVEs in dependencies", "0 critical, 0 high CVEs in dependency tree"),
]

# ──────────────────────────────────────────────
# Helper functions
# ──────────────────────────────────────────────

def _rand_int(lo, hi):
    return random.randint(lo, hi)


def _fill_template(template, module):
    """Replace placeholders in a template string with mock values."""
    s = template.replace("{module}", module)
    replacements = {
        "{t}": str(round(random.uniform(0.3, 2.8), 1)),
        "{w}": str(random.choice([375, 768, 1024, 1280, 1440, 1920])),
        "{p}": str(_rand_int(1, 20)),
        "{orient}": random.choice(["landscape", "portrait"]),
        "{fps}": str(_rand_int(55, 60)),
        "{scale}": str(round(random.uniform(1.0, 2.0), 1)),
        "{n}": str(random.choice([50, 100, 200, 300, 500, 1000])),
        "{rts}": str(random.choice([200, 300, 500])),
        "{rt}": str(_rand_int(40, 190)),
        "{p99}": str(_rand_int(150, 490)),
        "{p50}": str(_rand_int(10, 50)),
        "{p95}": str(_rand_int(80, 200)),
        "{p50r}": str(_rand_int(8, 45)),
        "{p95r}": str(_rand_int(70, 190)),
        "{p99r}": str(_rand_int(140, 480)),
        "{rec}": str(round(random.uniform(1.0, 4.5), 1)),
        "{mem}": str(round(random.uniform(0.5, 4.0), 1)),
        "{tps}": str(random.choice([500, 1000, 2000, 5000])),
        "{sz}": str(random.choice([64, 128, 256, 512, 1024])),
        "{rl}": str(random.choice([60, 100, 200, 500])),
        "{chr}": str(_rand_int(91, 99)),
        "{fo}": str(round(random.uniform(1.0, 2.9), 1)),
        "{comp}": str(_rand_int(62, 78)),
        "{reuse}": str(_rand_int(82, 97)),
        "{qt}": str(_rand_int(20, 100)),
        "{qtr}": str(_rand_int(15, 90)),
        "{th}": str(random.choice([70, 75, 80])),
        "{thr}": str(random.choice([70, 75, 80])),
        "{batch}": str(random.choice([5000, 10000, 20000, 50000])),
        "{to}": str(random.choice([5, 10, 15, 30])),
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    return s


def generate_test_cases(modules, actions, tc_prefix):
    """Return a list of 400 test-case dicts."""
    cases = []
    for i in range(1, NUM_CASES + 1):
        module = modules[(i - 1) % len(modules)]
        action_template = actions[(i - 1) % len(actions)]
        tc_name = _fill_template(action_template[0], module)
        expected = _fill_template(action_template[1], module)
        actual = _fill_template(action_template[2], module)
        mock_data = f"mock_user_{i:04d}, token_{random.randint(100000,999999)}, session_{random.randint(1000,9999)}"

        cases.append({
            "id": f"{tc_prefix}-{i:04d}",
            "name": tc_name,
            "description": f"Automated test: {tc_name}",
            "module": module,
            "test_data": mock_data,
            "expected": expected,
            "actual": actual,
            "status": "PASS",
            "remarks": "Executed successfully",
        })
    return cases


def write_report(filename, workflow_name, cases):
    """Create an Excel workbook with Summary + Detailed Report sheets."""
    wb = Workbook()

    # ── Summary Sheet ──────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "2F5496"

    # Title
    ws_summary.merge_cells("B2:E2")
    title_cell = ws_summary["B2"]
    title_cell.value = f"{workflow_name} – Execution Summary"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Summary data
    summary_rows = [
        ("Workflow Name", workflow_name),
        ("Total Test Cases Designed", NUM_CASES),
        ("Total Test Cases Executed", NUM_CASES),
        ("Total Test Cases Passed", NUM_CASES),
        ("Total Test Cases Failed", 0),
        ("Total Test Cases Skipped", 0),
        ("Overall Execution Status", "PASS"),
        ("Pass Percentage", "100%"),
        ("Execution Date & Time", EXECUTION_DATETIME),
    ]

    start_row = 4
    for idx, (label, value) in enumerate(summary_rows):
        row = start_row + idx
        lbl_cell = ws_summary.cell(row=row, column=2, value=label)
        val_cell = ws_summary.cell(row=row, column=3, value=value)
        lbl_cell.font = SUMMARY_LABEL_FONT
        lbl_cell.border = THIN_BORDER
        lbl_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        val_cell.font = SUMMARY_VALUE_FONT
        val_cell.border = THIN_BORDER
        val_cell.alignment = Alignment(horizontal="center")
        if label == "Overall Execution Status":
            val_cell.fill = PASS_FILL
            val_cell.font = PASS_FONT

    ws_summary.column_dimensions["B"].width = 32
    ws_summary.column_dimensions["C"].width = 28
    ws_summary.column_dimensions["D"].width = 5
    ws_summary.column_dimensions["E"].width = 5

    # ── Detailed Report Sheet ──────────────────
    ws_detail = wb.create_sheet("Detailed Report")
    ws_detail.sheet_properties.tabColor = "548235"

    headers = [
        "Test Case ID", "Test Case Name", "Test Case Description",
        "Module / Feature", "Test Data (Mock)", "Expected Result",
        "Actual Result", "Execution Status", "Remarks",
    ]
    col_widths = [16, 50, 55, 28, 45, 45, 45, 18, 24]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws_detail.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # Freeze top row
    ws_detail.freeze_panes = "A2"

    # Write data rows
    for row_idx, tc in enumerate(cases, start=2):
        values = [
            tc["id"], tc["name"], tc["description"],
            tc["module"], tc["test_data"], tc["expected"],
            tc["actual"], tc["status"], tc["remarks"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 8:  # Status column
                cell.fill = PASS_FILL
                cell.font = PASS_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-filter
    ws_detail.auto_filter.ref = f"A1:I{NUM_CASES + 1}"

    # Save
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"  [OK]  {filepath}")
    return filepath


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    random.seed(42)  # reproducibility

    workflows = [
        {
            "name": "Selenium Test",
            "file": "Selenium_Test_Report.xlsx",
            "prefix": "SEL",
            "modules": SELENIUM_MODULES,
            "actions": SELENIUM_ACTIONS,
        },
        {
            "name": "Appium Test",
            "file": "Appium_Test_Report.xlsx",
            "prefix": "APM",
            "modules": APPIUM_MODULES,
            "actions": APPIUM_ACTIONS,
        },
        {
            "name": "Load Test",
            "file": "Load_Test_Report.xlsx",
            "prefix": "LDT",
            "modules": LOAD_MODULES,
            "actions": LOAD_ACTIONS,
        },
        {
            "name": "Vulnerability Test",
            "file": "Vulnerability_Test_Report.xlsx",
            "prefix": "VLN",
            "modules": VULN_MODULES,
            "actions": VULN_ACTIONS,
        },
    ]

    print("Generating test reports...\n")
    for wf in workflows:
        cases = generate_test_cases(wf["modules"], wf["actions"], wf["prefix"])
        write_report(wf["file"], wf["name"], cases)

    print(f"\nDone – {len(workflows)} reports generated, {NUM_CASES} test cases each.")


if __name__ == "__main__":
    main()
